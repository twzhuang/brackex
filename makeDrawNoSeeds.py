from openpyxl import load_workbook, Workbook
from operator import itemgetter

def makeDraw():
    wb = load_workbook("masterlistspring.xlsx")

    for sheet in wb:
        if len(sheet.title) != 3:
            continue
        else:
            separatePlayersIntoFlights(sheet)


def separatePlayersIntoFlights(sheet):
    print(sheet.title)
    # possible flight combinations for each flight
    flightDict = {
        'A': ['A','AB'],
        'B': ['AB','B','BC'],
        'C': ['BC','C','CD'],
        'D': ['CD','D']
    }
    letter = sheet.title[0]
    # array of flights for the specific draw
    flights = flightDict[letter]

    # women's flights will be handled differently
    # call function to handle d women's flights
    # needs to be above all other singles and doubles

    # singles draws
    if sheet.title[2] == 'S':
        allFlightPlayers = separateSingles(flightDict, flights, sheet, letter)
    # doubles draws need to add partners
    else:
        allFlightPlayers = separateDoubles(flightDict, flights, sheet, letter)

    # sort players by club within each flight
    sortedByFlightByClub = sortClubs(allFlightPlayers)
    numRows = getNumRows(sortedByFlightByClub)

    # if numPlayers > 32 and numPlayers < 64:
    # number of matches that can be filled on the second round minus pullouts
    if numRows >= 64:
        oddIndex = [0,15,8,7,4,11,12,3,2,13,10,5,6,9,14,1]
        evenIndex =[15,0,7,8,11,4,3,12,13,2,5,10,9,6,1,14]
        smallerBracket = 64
    elif numRows >= 32 and numRows < 64:
        oddIndex = [0,7,4,3,5,2,6,1]
        evenIndex = [7,0,3,4,2,5,1,6]
        smallerBracket = 32
    elif numRows >= 16 and numRows < 32:
        oddIndex = [0,3,2,1]
        evenIndex = [3,0,1,2]
        smallerBracket = 16
    elif numRows >= 8 and numRows < 16:
        oddIndex = [0,1]
        evenIndex = [1,0]
        smallerBracket = 8
    else:
        print(sheet.title, "is too small to make a draw. You need at least 8 players to create a draw")
        return
    nonPullouts = smallerBracket - (numRows - smallerBracket)
    # print(nonPullouts)

    players = []
    for pList in sortedByFlightByClub:
        for player in pList:
            players.append(player)
            # print(player)


    # this will create an array representing our small bracket of the draw filled with players and empty arrays that represent pullout pulloutmatches
    bracketPlusIndexes = fillSmallBracket(oddIndex, evenIndex, smallerBracket, nonPullouts, players)
    filledBracket = bracketPlusIndexes[0]
    pulloutIndexes = bracketPlusIndexes[1]
    # print(filledBracket)
    pulloutIndexes += pulloutIndexes[::-1]
    # print(pulloutIndexes)

    completeDraw = fillSmallBracketWithPullouts(filledBracket, pulloutIndexes, players, nonPullouts)
    # print(completeDraw)
    # for match in completeDraw:
    #     print(match)

    printDraw(numRows, completeDraw, sheet.title)


def printDraw(numPlayers, draw, sheetName):
    drawTemplate = load_workbook("drawTemplate - Copy.xltx")
    sheets = drawTemplate.sheetnames
    if numPlayers == 8:
        source = drawTemplate[sheets[4]]
        sheet = drawTemplate.copy_worksheet(source)
    elif numPlayers <= 16:
        source = drawTemplate[sheets[3]]
        sheet = drawTemplate.copy_worksheet(source)
    elif numPlayers <= 32:
        source = drawTemplate[sheets[2]]
        sheet = drawTemplate.copy_worksheet(source)
    elif numPlayers <= 64:
        source = drawTemplate[sheets[1]]
        sheet = drawTemplate.copy_worksheet(source)
    elif numPlayers <= 128:
        source = drawTemplate[sheets[0]]
        sheet = drawTemplate.copy_worksheet(source)
    print(sheet.title)

    curRow = 8
    print(sheet[3][0].value)
    for player in draw:
        print(player)
        # inner match
        if sheetName[2] == 'S':
            if len(player) != 2:
                sheet[curRow][2].value = player[3] + " " + player[2] + " " + player[0] + " " + player[1]
            # pullout match
            else:
                sheet[curRow-3][0].value = player[0][3] + " " + player[0][2] + " " + player[0][0] + " " + player[0][1]
                sheet[curRow+2][0].value = player[1][3] + " " + player[1][2] + " " + player[1][0] + " " + player[1][1]
            curRow += 9
        else:
            if len(player) != 2:
                sheet[curRow][2].value = player[3] + " " + player[2] + " " + player[0] + " " + player[1] + " / " + player[4]
            # pullout match
            else:
                sheet[curRow-3][0].value = player[0][3] + " " + player[0][2] + " " + player[0][0] + " " + player[0][1] + " / " + player[0][4]
                sheet[curRow+2][0].value = player[1][3] + " " + player[1][2] + " " + player[1][0] + " " + player[1][1] + " / " + player[1][4]
            curRow += 9

    sheet.title = sheetName
    drawTemplate.save("drawTemplate - Copy.xltx")

def fillSmallBracketWithPullouts(draw, pulloutIndexes, players, start):
    j = 0
    for i in range(start, len(players)):
        draw[pulloutIndexes[j]].append(players[i])
        j += 1
    return draw
    

def fillSmallBracket(oddIndex, evenIndex, smallerBracket, nonPullouts, players):
    firstquadrant = [None]*int(smallerBracket/4)
    secondquadrant = [None]*int(smallerBracket/4)
    thirdquadrant = [None]*int(smallerBracket/4)
    fourthquadrant = [None]*int(smallerBracket/4)

    # when numNonPulloutsPlaced equals nonPullouts, we will start placing empty arrays in our bracket
    numNonPulloutsPlaced = 0
    count = 0
    # we only increment j after we've placed someone in all 4 quadrants of our bracket
    i = 0
    pulloutIndexes = []
    for j in range(smallerBracket):
        if (numNonPulloutsPlaced == nonPullouts):
            if count == 0:
                thirdquadrant[oddIndex[i]] = []
                count = 1
                pulloutIndex = oddIndex[i] + int(smallerBracket/4) * 2
            elif count == 1:
                secondquadrant[evenIndex[i]] = []
                count = 2
                pulloutIndex = evenIndex[i] + int(smallerBracket/4)
            elif count == 2:
                fourthquadrant[evenIndex[i]] = []
                count = 3
                pulloutIndex = evenIndex[i] + int(smallerBracket/4) * 3
            elif count == 3:
                firstquadrant[oddIndex[i]] = []
                count = 0
                pulloutIndex = oddIndex[i]  
                i += 1
            pulloutIndexes.append(pulloutIndex)

        else:
            if count == 0:
                thirdquadrant[oddIndex[i]] = players[j]
                count = 1
            elif count == 1:
                secondquadrant[evenIndex[i]] = players[j]
                count = 2
            elif count == 2:
                fourthquadrant[evenIndex[i]] = players[j]
                count = 3
            elif count == 3:
                firstquadrant[oddIndex[i]] = players[j]
                count = 0
                i += 1
            numNonPulloutsPlaced += 1
    draw = firstquadrant + secondquadrant + thirdquadrant + fourthquadrant
    return [draw, pulloutIndexes]


# get total number of players/pairs for that draw
def getNumRows(players):
    count = 0
    for playerList in players:
        for player in playerList:
            count += 1
    return count

# function to sort each array in given array by club
# returns sorted array separated by flight
def sortClubs(players):
    allPlayersSorted = []
    for listPlayers in players:
        flightList = sorted(listPlayers, key=itemgetter(2), reverse = False)
        allPlayersSorted.append(flightList)
    return allPlayersSorted

def separateSingles(flightDict, flights, sheet, letter):
    lfplayers = []
    mfplayers = []
    hfplayers = []
    for row in sheet:
        if row[0].value == 'Last Name':
            continue
        ln = row[0].value
        fn = row[1].value
        club = row[4].value
        flight = row[3].value
        if club == None:
            club='Z'


        if row[3].value == flights[0]:
            hfplayers.append([fn,ln,club,flight])
        if row[3].value == flights[1]:
            mfplayers.append([fn,ln,club,flight])
        # if the flight is B or C, need to add a low flight
        if letter == 'B' or letter == 'C':
            if row[3].value == flights[2]:
                lfplayers.append([fn,ln,club,flight])


    if hfplayers == []:
        return [mfplayers, lfplayers]
    elif lfplayers == []:
        return [hfplayers, mfplayers]
    else:
        return [hfplayers, mfplayers, lfplayers]

def separateDoubles(flightDict, flights, sheet, letter):
    lfplayers = []
    mfplayers = []
    hfplayers = []
    for row in sheet:
        if row[0].value == 'Last Name':
            continue
        ln = row[0].value
        fn = row[1].value
        club = row[5].value
        flight = row[3].value
        partner = row[4].value
        if club == None:
            club = 'Z'

        if row[3].value == flights[0]:
            hfplayers.append([fn,ln,club,flight,partner])
        if row[3].value == flights[1]:
            mfplayers.append([fn,ln,club,flight,partner])
        if letter == 'B' or letter == 'C':
            if row[3].value == flights[2]:
                lfplayers.append([fn,ln,club,flight,partner])
    if hfplayers == []:
        return [mfplayers, lfplayers]
    elif lfplayers == []:
        return [hfplayers, mfplayers]
    else:
        return [hfplayers, mfplayers, lfplayers]



def main():
    makeDraw()

if __name__ == '__main__':
    main()
