from openpyxl import load_workbook, Workbook
from operator import itemgetter

def makeDraw():
    wb = load_workbook("playerParseEdit.xlsx")

    for sheet in wb:
        if len(sheet.title) != 3:
            continue
        else:
            separatePlayersIntoFlights(sheet)


def separatePlayersIntoFlights(sheet):
    print(sheet.title)
    # possible flight combinations for each flight
    flightDict = {
        'A': ['A','AB'],
        'B': ['AB','B','BC'],
        'C': ['BC','C','CD'],
        'D': ['CD','D']
    }
    letter = sheet.title[0]
    # array of flights for the specific draw
    flights = flightDict[letter]

    # women's flights will be handled differently
    # call function to handle d women's flights
    # needs to be above all other singles and doubles

    # singles draws
    if sheet.title[2] == 'S':
        allFlightPlayers = separateSingles(flightDict, flights, sheet, letter)
        numSeeds = len(allFlightPlayers[0])
        sortedByFlightByClub = sortClubs(allFlightPlayers)
        sortedByFlightByClub[0] = sorted(sortedByFlightByClub[0], key=itemgetter(4), reverse=False)
    # doubles draws need to add partners
    else:
        allFlightPlayers = separateDoubles(flightDict, flights, sheet, letter)
        numSeeds = len(allFlightPlayers[0])
        sortedByFlightByClub = sortClubs(allFlightPlayers)
        sortedByFlightByClub[0] = sorted(sortedByFlightByClub[0], key=itemgetter(5), reverse=False)
    # sort players by club within each flight

    seededPlayers = sortedByFlightByClub[0]
    print("Seeded players:",seededPlayers)
    if len(seededPlayers) >= 4:
        seededPlayers[3], seededPlayers[2] = seededPlayers[2], seededPlayers[3]
    if len(seededPlayers) == 8:
        seededPlayers[4], seededPlayers[5] = seededPlayers[5], seededPlayers[4]
    print("Seeded players:",seededPlayers)

    print("number of seeds:", numSeeds)
    print(sortedByFlightByClub)
    
    numRows = getNumRows(sortedByFlightByClub)

    # if numPlayers > 32 and numPlayers < 64:
    # number of matches that can be filled on the second round minus pullouts
    if numRows >= 64:
        oddIndex = [0,15,8,7,4,11,12,3,2,13,10,5,6,9,14,1]
        evenIndex =[15,0,7,8,11,4,3,12,13,2,5,10,9,6,1,14]
        smallerBracket = 64
    elif numRows >= 32 and numRows < 64:
        oddIndex = [0,7,4,3,5,2,6,1]
        evenIndex = [7,0,3,4,2,5,1,6]
        smallerBracket = 32
    elif numRows >= 16 and numRows < 32:
        oddIndex = [0,3,2,1]
        evenIndex = [3,0,1,2]
        smallerBracket = 16
    elif numRows >= 8 and numRows < 16:
        oddIndex = [0,1]
        evenIndex = [1,0]
        smallerBracket = 8
    else:
        print(sheet.title, "is too small to make a draw. You need at least 8 players to create a draw")
        return
    nonPullouts = smallerBracket - (numRows - smallerBracket)
    # print(nonPullouts)

    players = []
    for pList in sortedByFlightByClub:
        for player in pList:
            players.append(player)
            # print(player)

    # place seeds in draw first
    bracketWithSeeds = placeSeeds(oddIndex, evenIndex, smallerBracket, nonPullouts, players, numSeeds)



    # this will create an array representing our small bracket of the draw filled with players and empty arrays that represent pullout pulloutmatches
    bracketPlusIndexes = fillSmallBracket(bracketWithSeeds, numSeeds, oddIndex, evenIndex, smallerBracket, nonPullouts, players)
    filledBracket = bracketPlusIndexes[0]
    pulloutIndexes = bracketPlusIndexes[1]
    # print(filledBracket)
    pulloutIndexes += pulloutIndexes[::-1]
    # print(pulloutIndexes)

    completeDraw = fillSmallBracketWithPullouts(filledBracket, pulloutIndexes, players, nonPullouts)
    # print(completeDraw)
    # for match in completeDraw:
    #     print(match)

    printDraw(numRows, completeDraw, sheet.title)


def printDraw(numPlayers, draw, sheetName):
    drawTemplate = load_workbook("drawTemplate - Copy.xltx")
    sheets = drawTemplate.sheetnames
    if numPlayers == 8:
        source = drawTemplate[sheets[4]]
        sheet = drawTemplate.copy_worksheet(source)
    elif numPlayers <= 16:
        source = drawTemplate[sheets[3]]
        sheet = drawTemplate.copy_worksheet(source)
    elif numPlayers <= 32:
        source = drawTemplate[sheets[2]]
        sheet = drawTemplate.copy_worksheet(source)
    elif numPlayers <= 64:
        source = drawTemplate[sheets[1]]
        sheet = drawTemplate.copy_worksheet(source)
    elif numPlayers <= 128:
        source = drawTemplate[sheets[0]]
        sheet = drawTemplate.copy_worksheet(source)
    print(sheet.title)
    print("sheet name:", sheetName)

    sheet[1][0].value = "UC DAVIS BADMINTON FALL 2019 OPEN"
    sheet[3][0].value = sheetName

    print(sheet[3][0].value)

    # if draw is perfect draw, print everything in first column
    if numPlayers == 8 or numPlayers == 16 or numPlayers == 32 or numPlayers == 64:
        curRow = 5
        i = 0
        for player in draw:
            print("player:",player)
            if sheetName[2] == "S":
                sheet[curRow][0].value = player[3] + " " + player[2] + " " + player[0] + " " + player[1]
            else:
                sheet[curRow][0].value = player[3] + " " + player[2] + " " + player[0] + " " + player[1] + " / " + player[4]
            i += 1
            if i % 2 == 0:
                curRow += 4
            else: 
                curRow += 5

    else:
        curRow = 8
        for player in draw:
            print("player:",player)
            # inner match

            if sheetName[2] == 'S':
                if len(player) != 2:
                    sheet[curRow][2].value = player[3] + " " + player[2] + " " + player[0] + " " + player[1]
                # pullout match
                else:
                    sheet[curRow-3][0].value = player[0][3] + " " + player[0][2] + " " + player[0][0] + " " + player[0][1]
                    sheet[curRow+2][0].value = player[1][3] + " " + player[1][2] + " " + player[1][0] + " " + player[1][1]
                curRow += 9
            else:
                if len(player) != 2:
                    sheet[curRow][2].value = player[3] + " " + player[2] + " " + player[0] + " " + player[1] + " / " + player[4]
                # pullout match
                else:
                    sheet[curRow-3][0].value = player[0][3] + " " + player[0][2] + " " + player[0][0] + " " + player[0][1] + " / " + player[0][4]
                    sheet[curRow+2][0].value = player[1][3] + " " + player[1][2] + " " + player[1][0] + " " + player[1][1] + " / " + player[1][4]
                curRow += 9

    sheet.title = sheetName
    drawTemplate.save("drawTemplate - Copy.xltx")

def fillSmallBracketWithPullouts(draw, pulloutIndexes, players, start):
    j = 0
    for i in range(start, len(players)):
        draw[pulloutIndexes[j]].append(players[i])
        j += 1
    return draw
    

def fillSmallBracket(bracketWithSeeds, numSeeds, oddIndex, evenIndex, smallerBracket, nonPullouts, players):
    # program will place all non-pullout matches first
    # once number of non pullouts is reached, program will start placing empty [] in place of pullout matches 
    firstquadrant = bracketWithSeeds[0]
    secondquadrant = bracketWithSeeds[1]
    thirdquadrant = bracketWithSeeds[2]
    fourthquadrant = bracketWithSeeds[3]


    # when numNonPulloutsPlaced equals nonPullouts, we will start placing empty arrays in our bracket
    numNonPulloutsPlaced = bracketWithSeeds[5]
    count = 0
    # we only increment j after we've placed someone in all 4 quadrants of our bracket
    i = bracketWithSeeds[6]
    pulloutIndexes = bracketWithSeeds[4]
    for j in range(numSeeds,smallerBracket):
        # place pullout matches 
        if (numNonPulloutsPlaced >= nonPullouts):
            if count == 0:
                firstquadrant[oddIndex[i]] = []
                count = 1
                pulloutIndex = oddIndex[i]
            elif count == 1:
                fourthquadrant[evenIndex[i]] = []
                count = 2
                pulloutIndex = evenIndex[i] + int(smallerBracket/4) * 3
            elif count == 2:
                secondquadrant[evenIndex[i]] = []
                count = 3
                pulloutIndex = evenIndex[i] + int(smallerBracket/4)
            elif count == 3:
                thirdquadrant[oddIndex[i]] = []
                count = 0
                pulloutIndex = oddIndex[i] + int(smallerBracket/4) * 2
                i += 1
            pulloutIndexes.append(pulloutIndex)
        # if (numNonPulloutsPlaced == nonPullouts):
        #     if count == 0:
        #         thirdquadrant[oddIndex[i]] = []
        #         count = 1
        #         pulloutIndex = oddIndex[i] + int(smallerBracket/4) * 2
        #     elif count == 1:
        #         secondquadrant[evenIndex[i]] = []
        #         count = 2
        #         pulloutIndex = evenIndex[i] + int(smallerBracket/4)
        #     elif count == 2:
        #         fourthquadrant[evenIndex[i]] = []
        #         count = 3
        #         pulloutIndex = evenIndex[i] + int(smallerBracket/4) * 3
        #     elif count == 3:
        #         firstquadrant[oddIndex[i]] = []
        #         count = 0
        #         pulloutIndex = oddIndex[i]  
        #         i += 1
        #     pulloutIndexes.append(pulloutIndex)

        # place non-pullouts
        else:
            if count == 0:
                firstquadrant[oddIndex[i]] = players[j]
                count = 1
            elif count == 1:
                fourthquadrant[evenIndex[i]] = players[j]
                count = 2
            elif count == 2:
                secondquadrant[evenIndex[i]] = players[j]
                count = 3
            elif count == 3:
                thirdquadrant[oddIndex[i]] = players[j]
                count = 0
                i += 1
            numNonPulloutsPlaced += 1
            # if count == 0:
            #     thirdquadrant[oddIndex[i]] = players[j]
            #     count = 1
            # elif count == 1:
            #     secondquadrant[evenIndex[i]] = players[j]
            #     count = 2
            # elif count == 2:
            #     fourthquadrant[evenIndex[i]] = players[j]
            #     count = 3
            # elif count == 3:
            #     firstquadrant[oddIndex[i]] = players[j]
            #     count = 0
            #     i += 1
            # numNonPulloutsPlaced += 1
    draw = firstquadrant + secondquadrant + thirdquadrant + fourthquadrant
    return [draw, pulloutIndexes]

def placeSeeds(oddIndex, evenIndex, smallerBracket, nonPullouts, players, numSeeds):
    firstquadrant = [None]*int(smallerBracket/4)
    secondquadrant = [None]*int(smallerBracket/4)
    thirdquadrant = [None]*int(smallerBracket/4)
    fourthquadrant = [None]*int(smallerBracket/4)


    # when numNonPulloutsPlaced equals nonPullouts, we will start placing empty arrays in our bracket
    numNonPulloutsPlaced = 0
    count = 0
    # we only increment j after we've placed someone in all 4 quadrants of our bracket
    i = 0
    pulloutIndexes = []
    # smaller bracket is the number of players on inside 
    for j in range(numSeeds):
        # place pullout matches 
        if (numNonPulloutsPlaced == nonPullouts):
            if count == 0:
                firstquadrant[oddIndex[i]] = []
                count = 1
                pulloutIndex = oddIndex[i]
            elif count == 1:
                fourthquadrant[evenIndex[i]] = []
                count = 2
                pulloutIndex = evenIndex[i] + int(smallerBracket/4) * 3
            elif count == 2:
                secondquadrant[evenIndex[i]] = []
                count = 3
                pulloutIndex = evenIndex[i] + int(smallerBracket/4)
            elif count == 3:
                thirdquadrant[oddIndex[i]] = []
                count = 4
                pulloutIndex = oddIndex[i] + int(smallerBracket/4) * 2
                i += 1
            elif count == 4:
                thirdquadrant[oddIndex[i]] = []
                count = 5
                pulloutIndex = oddIndex[i] + int(smallerBracket/4) * 2
            elif count == 5:
                secondquadrant[evenIndex[i]] = []
                count = 6
                pulloutIndex = evenIndex[i] + int(smallerBracket/4)
            elif count == 6:
                fourthquadrant[evenIndex[i]] = []
                count = 7
                pulloutIndex = evenIndex[i] + int(smallerBracket/4) * 3
            elif count == 7:
                firstquadrant[oddIndex[i]] = []
                count = 0
                pulloutIndex = oddIndex[i]
                i += 1
            print(count)
            pulloutIndexes.append(pulloutIndex)

        # place seeds up to 8
        else:
            if count == 0:
                firstquadrant[oddIndex[i]] = players[j]
                count = 1
            elif count == 1:
                fourthquadrant[evenIndex[i]] = players[j]
                count = 2
            elif count == 2:
                secondquadrant[evenIndex[i]] = players[j]
                count = 3
            elif count == 3:
                thirdquadrant[oddIndex[i]] = players[j]
                count = 4
                i += 1
            elif count == 4:
                thirdquadrant[oddIndex[i]] = players[j]
                count = 5
            elif count == 5:
                secondquadrant[evenIndex[i]] = players[j]
                count = 6
            elif count == 6:
                fourthquadrant[evenIndex[i]] = players[j]
                count = 7
            elif count == 7:
                firstquadrant[oddIndex[i]] = players[j]
                count = 0
                i += 1
            numNonPulloutsPlaced += 1
    # draw = firstquadrant + secondquadrant + thirdquadrant + fourthquadrant
    return [firstquadrant, secondquadrant, thirdquadrant, fourthquadrant, pulloutIndexes, numNonPulloutsPlaced, i]

# get total number of players/pairs for that draw
def getNumRows(players):
    count = 0
    for playerList in players:
        for player in playerList:
            count += 1
    return count

# function to sort each array in given array by club
# returns sorted array separated by flight
def sortClubs(players):
    allPlayersSorted = []
    for listPlayers in players:
        flightList = sorted(listPlayers, key=itemgetter(2), reverse = False)
        allPlayersSorted.append(flightList)
    return allPlayersSorted

def separateSingles(flightDict, flights, sheet, letter):
    lfplayers = []
    mfplayers = []
    hfplayers = []
    seededPlayers = []
    for row in sheet:
        if row[0].value == 'Last Name':
            continue
        ln = row[0].value
        fn = row[1].value
        club = row[4].value
        seed = row[5].value
        flight = row[3].value
        if club == None:
            club='Z'

        if seed != None:
            seededPlayers.append([fn, ln, club, flight, seed])

        if row[3].value == flights[0] and seed == None:
            hfplayers.append([fn,ln,club,flight])
        if row[3].value == flights[1] and seed == None:
            mfplayers.append([fn,ln,club,flight])
        # if the flight is B or C, need to add a low flight
        if letter == 'B' or letter == 'C':
            if row[3].value == flights[2] and seed == None:
                lfplayers.append([fn,ln,club,flight])
    if hfplayers == []:
        return [seededPlayers, mfplayers, lfplayers]
    elif lfplayers == []:
        return [seededPlayers, hfplayers, mfplayers]
    else:
        return [seededPlayers, hfplayers, mfplayers, lfplayers]

def separateDoubles(flightDict, flights, sheet, letter):
    lfplayers = []
    mfplayers = []
    hfplayers = []
    seededPlayers = []
    for row in sheet:
        if row[0].value == 'Last Name':
            continue
        ln = row[0].value
        fn = row[1].value
        club = row[5].value
        seed = row[6].value
        flight = row[3].value
        partner = row[4].value
        if club == None:
            club = 'Z'

        if seed != None:
            seededPlayers.append([fn,ln,club,flight,partner,seed])
        if row[3].value == flights[0] and seed == None:
            hfplayers.append([fn,ln,club,flight,partner])
        if row[3].value == flights[1] and seed == None:
            mfplayers.append([fn,ln,club,flight,partner])
        if letter == 'B' or letter == 'C':
            if row[3].value == flights[2] and seed == None:
                lfplayers.append([fn,ln,club,flight,partner])

    if hfplayers == []:
        return [seededPlayers, mfplayers, lfplayers]
    elif lfplayers == []:
        return [seededPlayers, hfplayers, mfplayers]
    else:
        return [seededPlayers, hfplayers, mfplayers, lfplayers]



def main():
    makeDraw()

if __name__ == '__main__':
    main()
