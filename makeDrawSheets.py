from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font 

import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile

def parseFile():
    wb = load_workbook('masterlistEdit.xlsx')
    main = wb['Players']

    flights = {
        'AMD': 0, 'BMD': 0,'CMD': 0,'DMD': 0,'AWD': 0,
        'BWD': 0,'CWD': 0,'DWD': 0,'AMX': 0,'BMX': 0,
        'CMX': 0,'DMX': 0,'AMS': 0,'BMS': 0,'CMS': 0,
        'DMS': 0,'AWS': 0,'BWS': 0,'CWS': 0,'DWS': 0,
        'AXD': 0,'BXD': 0,'CXD': 0,'DXD': 0
    }

    header = list(main.rows)[0]
    print(header)
    # for cell in header:
    #     cell.font = Font(bold=True)

    # loop through every row and create a new sheet for each event
    # player will be appended to the event's sheet
    for row in main.rows:
        # print(row[11].value)
        events = row[11].value
        bottomBorder = Border(bottom=Side(style='thin'))
        # skip first couple lines
        if (events != None):
            eventList = events.split(", ")
            # print(eventList)
            for event in eventList:
                if event in flights:
                    if flights[event] == 0:
                        flights[event] += 1
                        newSheet = wb.create_sheet(event)
                        newSheet.append(cell.value for cell in header)
                        for cell in newSheet["1:1"]:
                            cell.font = Font(bold = True)
                            cell.border = bottomBorder
                        newSheet.append(cell.value for cell in row)
                    else:
                        wb[event].append(cell.value for cell in row)

    deleteColumns(wb)
    deletePartners(wb)

    # sort sheets in alphabetical order
    wb._sheets.sort(key=lambda ws:ws.title)
    wb.save('playerParse.xlsx')

    sortSheets()

def deletePartners(wb):
    f = open("playersToDelete.txt", "w+")
    for sheet in wb:
        # if doubles or mixed sheet, need to delete partners off of sheet
        if sheet.title[-1] == "D" or sheet.title[-1] == "X":
            players = {}
            rowsToDelete = []
            for row in sheet:
                # skip first row
                if row[0].value == "Last Name":
                    continue
                playerName = row[1].value + " " + row[0].value
                # if player's partner was already stored in players dictionary,
                # mark row for deletion
                partnerName = row[4].value
                if partnerName in players:
                    rowsToDelete.append(players[row[4].value])
                    del players[partnerName]
                else:
                    players[playerName] = row[0].row

            # delete rows marked for deletion
            for row in sorted(rowsToDelete, reverse=True):
                print("deleting row", row)
                sheet.delete_rows(row)
            
            # if a player from each pair wasn't deleted, remaining players will be written down
            f.write("******************" + sheet.title + "********************\n")
            for player in players:
                f.write(player + " still needs to be removed from list\n")
            print(players)
        print(sheet.title)

def sortSheets():
    xl = pd.ExcelFile("playerParse.xlsx")
    dfDict = {}
    for sheet in xl.sheet_names:
        df = xl.parse(sheet)
        df = df.sort_values(["Flights"], ascending=True)
        dfDict[sheet] = df
    writer = pd.ExcelWriter("playerParse.xlsx")

    for sheetTitle in dfDict:
        dfDict[sheetTitle].to_excel(writer, sheetTitle, index=False)
    writer.save()


def deleteColumns(wb):
    for sheet in wb:
        # remove rows for corresponding event
        sheet.delete_cols(12,2)
        if sheet.title[1:] == "MS":
            sheet.delete_cols(5,6)
        elif sheet.title[1:] == "WS":
            sheet.delete_cols(6,5)
            sheet.delete_cols(4,1)
        elif sheet.title[1:] == "MX":
            sheet.delete_cols(10,1)
            sheet.delete_cols(7,2)
            sheet.delete_cols(4,2)
        elif sheet.title[1:] == "MD":
            sheet.delete_cols(8,2)
            sheet.delete_cols(4,3)
        elif sheet.title[1:] == "WD":
            sheet.delete_cols(9,1)
            sheet.delete_cols(4,4)
        
        if sheet.title[-1] == "D" or sheet.title[-1] == "X":
             sheet[1][4].value = "Partner"
        
        sheet[1][3].value = "Flights"

def main():
    parseFile()

if __name__ == "__main__":
    main()
